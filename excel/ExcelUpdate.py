import pandas as pd
from datetime import date,datetime
from Node import Node
from pandas import DataFrame
import sys
import xlrd
from openpyxl import load_workbook

class Solution:
    def generate(self, filePath_source: str, filePath_Destination: str):
        data_source = pd.read_excel(filePath_source)
        df_source = pd.DataFrame(data_source, columns= ['Plate Name','Inspection Date Time', 'Defects'])
        for ind in df_source.index:
            plate_name = df_source['Plate Name'][ind]
            inspection_date_time = df_source['Inspection Date Time'][ind]
            Defects =  df_source['Defects'][ind]
            node = Node(plate_name, inspection_date_time, Defects)
            self.updateDestination(filePath_Destination, node)

    def updateDestination(self, file_home, node):
        wb = load_workbook(filename= file_home)
        sheet_ranges = wb['Daily STL Check']
        index = 111
        while index <= 500:
            reticleId = sheet_ranges.cell(row=index, column=5).value
            if reticleId:
                reticleId = reticleId.replace('-', '')
                reticleId = reticleId[:-2]
                if reticleId == node.plate_name:
                    inspection_date_time = node.inspection_date_time.strftime('%Y/%m/%d')
                    sheet_ranges["P"+ str(index)] = inspection_date_time
                    sheet_ranges["Q" + str(index)] = node.defects_number
            index += 1
        wb.save(file_home)

def main(argv):
    source = argv[1]
    destination = argv[2]
    print(source)
    print(destination)
    sol = Solution()
    result = sol.generate(source, destination)
    print("Done")

if __name__ == "__main__":
    main(sys.argv)