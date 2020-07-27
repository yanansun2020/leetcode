from openpyxl import load_workbook
from Node import Node
class Solution:
    def test1(self, node):
        file_home = '/home/yanan/work/leetcode/excel/target table.xlsx'
        wb = load_workbook(filename= file_home)
        sheet_ranges = wb['Daily STL Check']
        index = 111
        while index <= 200:
            reticleId = sheet_ranges.cell(row=index, column=5).value
            if reticleId:
                reticleId = reticleId.replace('-', '')
                reticleId = reticleId[:-2]
                if reticleId == node.plate_name:
                    inspection_date_time = node.inspection_date_time.split(" ")[0]
                    sheet_ranges["P"+ str(index)] = inspection_date_time
                    sheet_ranges["Q" + str(index)] = node.defects_number
            index += 1
        wb.save(file_home)
def main():
    sol = Solution()
    node = Node("Y00100156AA1", "6/3/2020 16:08", "22")
    sol.test1(node)

if __name__ == "__main__":
    main()